import warnings
warnings.simplefilter(action='ignore')

import os
import timeit
import duckdb
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from datetime import datetime

class DuckDBConnector:
    def __init__(self, db=':memory:'):
        self.db = db
        self.conn = None
    
    def connect_to_db(self):
        """Connects to a DuckDB database and installs spatial extension."""
        self.conn = duckdb.connect(self.db)
        self.conn.install_extension('spatial')
        self.conn.load_extension('spatial')
        return self.conn
    
    def disconnect_db(self):
        """Disconnects from the DuckDB database."""
        if self.conn is not None:
            self.conn.close()
            self.conn = None
 

def write_dfs_to_excel(dataframes, filename):
    # Create a new Excel workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active

    current_row = 1
    
    for idx, df in enumerate(dataframes):
        # Round all numeric values to 1 decimal place
        df = df.round(2)

        # Write the dataframe to the Excel sheet, ignoring the index
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=current_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Define table range
        table_ref = f"A{current_row}:"
        last_column_letter = chr(64 + len(df.columns))  # Get last column's letter
        table_ref += f"{last_column_letter}{current_row + len(df)}"
        
        # Add table layout with a default style
        table = Table(displayName=f"Table{idx+1}", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Move to the next block of rows for the next dataframe (4-row space)
        current_row += len(df) + 4  # 1 row for the header + df rows + 4 empty rows
    
    # Set column widths: first 4 columns to 14, the rest to 24.5
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 14 if col_idx <= 4 else 24.5
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Save the workbook
    wb.save(filename)
    

def export_dfs_to_sheets(dataframes, sheet_names, filename):
    # Create a new Excel workbook
    wb = Workbook()

    # Remove the default sheet created with the workbook
    wb.remove(wb.active)

    for df, sheet_name in zip(dataframes, sheet_names):
        # Add a new sheet for each dataframe
        ws = wb.create_sheet(title=sheet_name)

        # Write the dataframe to the Excel sheet, ignoring the index
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(filename)  
            

if __name__ == "__main__":
    start_t = timeit.default_timer() #start time 
    
    wks= r'W:\lwbc\visr\Workarea\moez_labiadh\WORKSPACE_2024\20240819_flp_to_thlb_analysis'

    print ('Connecting to databases')    
    print ('..connect to Duckdb') 
    projDB= os.path.join(wks, 'inputs', 'tor_flp_thlb_analysis.db')
    Duckdb= DuckDBConnector(db= projDB)
    Duckdb.connect_to_db()
    dckCnx= Duckdb.conn 
    dckCnx.execute("SET GLOBAL pandas_analyze_sample=1000000")
    
    try:

       
        print ('\nCompute RIP/OGDA summary')

        df_rpog= dckCnx.execute("""SELECT* FROM r2_2_rip_ogda_thlb""").df()
        
        df_rpog['THLB_AREA_DECREASE']= df_rpog['AREA_HA'] * df_rpog['thlb_fact']
        
        df_rpog_sum = df_rpog.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()

        

        print ('\nCompute RIP/IDF summaries')
        
        df_rpdf= dckCnx.execute("""SELECT* FROM r2_2_rip_idf_thlb_mdwr""").df()
        
        df_rpdf= df_rpdf.rename(columns={'TSA_NUMBER_DESCRIPTION': 'TSA_NAME'})
        
        df_rpdf['THLB_AREA']= df_rpdf['AREA_HA'] * df_rpdf['thlb_fact']
        

        ####### Okanagan scenarios #######
        df_rpdf_okn= df_rpdf[df_rpdf['TSA_NAME']=='Okanagan TSA']
        
        df_rpdf_okn = df_rpdf_okn[
            ~df_rpdf_okn['BEC_SUBZONE'].isin(['mw', 'mm']) | df_rpdf_okn['BEC_SUBZONE'].isnull()
        ]
        
            ## scenario-1 ##
        df_rpdf_okn_s1 = df_rpdf_okn[
            (df_rpdf_okn['PROJ_AGE_1'] >= 100) | df_rpdf_okn['PROJ_AGE_1'].isnull()
        ]
        
        
        df_rpdf_okn_s1['REDUCTION_FACTOR_S1'] = np.where(
            df_rpdf_okn_s1['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdf_okn_s1['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)
        )

        
        df_rpdf_okn_s1['THLB_AREA_DECREASE'] = df_rpdf_okn_s1['THLB_AREA'] * df_rpdf_okn_s1['REDUCTION_FACTOR_S1']
        
        df_rpdf_okn_s1_sum = df_rpdf_okn_s1.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdf_okn_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        


            ## scenario-2 ##
        df_rpdf_okn_s2 = df_rpdf_okn[
            (df_rpdf_okn['PROJ_AGE_1'] >= 60) | df_rpdf_okn['PROJ_AGE_1'].isnull()
        ]
        
        
        df_rpdf_okn_s2['REDUCTION_FACTOR_S2'] = np.where(
            df_rpdf_okn_s2['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdf_okn_s2['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)
        )

        
        df_rpdf_okn_s2['THLB_AREA_DECREASE'] = df_rpdf_okn_s2['THLB_AREA'] * df_rpdf_okn_s2['REDUCTION_FACTOR_S2']
        
        df_rpdf_okn_s2_sum = df_rpdf_okn_s2.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdf_okn_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_rpdf_okn_fn = pd.concat([df_rpdf_okn_s1_sum, df_rpdf_okn_s2_sum], ignore_index=True)
        

          
        
        ####### Kamloops scenarios #######
        df_rpdf_kam= df_rpdf[df_rpdf['TSA_NAME']=='Kamloops TSA']
        
        df_rpdf_kam = df_rpdf_kam[
            ~df_rpdf_kam['BEC_SUBZONE'].isin(['mw', 'mm']) | df_rpdf_kam['BEC_SUBZONE'].isnull()
        ]
        
            ## scenario-1 ##
        df_rpdf_kam_s1 = df_rpdf_kam[
            (df_rpdf_kam['PROJ_AGE_1'] >= 100) | df_rpdf_kam['PROJ_AGE_1'].isnull()
        ]
        
        

        df_rpdf_kam_s1['REDUCTION_FACTOR_S1'] = np.where(
            df_rpdf_kam_s1['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdf_kam_s1['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        )      
        
        df_rpdf_kam_s1['THLB_AREA_DECREASE'] = df_rpdf_kam_s1['THLB_AREA'] * df_rpdf_kam_s1['REDUCTION_FACTOR_S1']
        
        df_rpdf_kam_s1_sum = df_rpdf_kam_s1.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdf_kam_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')


      

            ## scenario-2 ##
        df_rpdf_kam_s2 = df_rpdf_kam[
            (df_rpdf_kam['PROJ_AGE_1'] >= 60) | df_rpdf_kam['PROJ_AGE_1'].isnull()
        ]
        
        

        df_rpdf_kam_s2['REDUCTION_FACTOR_S2'] = np.where(
            df_rpdf_kam_s2['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdf_kam_s2['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        )      
        
        df_rpdf_kam_s2['THLB_AREA_DECREASE'] = df_rpdf_kam_s2['THLB_AREA'] * df_rpdf_kam_s2['REDUCTION_FACTOR_S2']
        
        df_rpdf_kam_s2_sum = df_rpdf_kam_s2.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdf_kam_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')


        df_rpdf_kam_fn = pd.concat([df_rpdf_kam_s1_sum, df_rpdf_kam_s2_sum], ignore_index=True)
        
        df_rpdf_fnl = pd.concat([df_rpdf_okn_fn, df_rpdf_kam_fn], ignore_index=True)

        
 
        print ('\nCompute RIP/IDF/OGDA summaries')
        
        df_rpdfog= dckCnx.execute("""SELECT* FROM r2_2_rip_idf_ogda_thlb_mdwr""").df()
        
        df_rpdfog= df_rpdfog.rename(columns={'TSA_NUMBER_DESCRIPTION': 'TSA_NAME'})
        
        df_rpdfog['THLB_AREA']= df_rpdfog['AREA_HA'] * df_rpdfog['thlb_fact']
        
  

        ####### Okanagan scenarios #######
        df_rpdfog_okn= df_rpdfog[df_rpdfog['TSA_NAME']=='Okanagan TSA']
        
        df_rpdfog_okn = df_rpdfog_okn[
            ~df_rpdfog_okn['BEC_SUBZONE'].isin(['mw', 'mm']) | df_rpdfog_okn['BEC_SUBZONE'].isnull()
        ]
        
            ## scenario-1 ##
        df_rpdfog_okn_s1 = df_rpdfog_okn[
            (df_rpdfog_okn['PROJ_AGE_1'] >= 100) | df_rpdfog_okn['PROJ_AGE_1'].isnull()
        ]
        
        
        df_rpdfog_okn_s1['REDUCTION_FACTOR_S1'] = np.where(
            df_rpdfog_okn_s1['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdfog_okn_s1['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)
        )

        
        df_rpdfog_okn_s1['THLB_AREA_DECREASE'] = df_rpdfog_okn_s1['THLB_AREA'] * df_rpdfog_okn_s1['REDUCTION_FACTOR_S1']
        
        df_rpdfog_okn_s1_sum = df_rpdfog_okn_s1.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdfog_okn_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        

            ## scenario-2 ##
        df_rpdfog_okn_s2 = df_rpdfog_okn[
            (df_rpdfog_okn['PROJ_AGE_1'] >= 60) | df_rpdfog_okn['PROJ_AGE_1'].isnull()
        ]
        
        
        df_rpdfog_okn_s2['REDUCTION_FACTOR_S2'] = np.where(
            df_rpdfog_okn_s2['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdfog_okn_s2['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)
        )

        
        df_rpdfog_okn_s2['THLB_AREA_DECREASE'] = df_rpdfog_okn_s2['THLB_AREA'] * df_rpdfog_okn_s2['REDUCTION_FACTOR_S2']
        
        df_rpdfog_okn_s2_sum = df_rpdfog_okn_s2.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdfog_okn_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_rpdfog_okn_fn = pd.concat([df_rpdfog_okn_s1_sum, df_rpdfog_okn_s2_sum], ignore_index=True)
        
        
        ####### Kamloops scenarios #######
        df_rpdfog_kam= df_rpdfog[df_rpdfog['TSA_NAME']=='Kamloops TSA']
        
        df_rpdfog_kam = df_rpdfog_kam[
            ~df_rpdfog_kam['BEC_SUBZONE'].isin(['mw', 'mm']) | df_rpdfog_kam['BEC_SUBZONE'].isnull()
        ]  
        
            ## scenario-1 ##
        df_rpdfog_kam_s1 = df_rpdfog_kam[
            (df_rpdfog_kam['PROJ_AGE_1'] >= 100) | df_rpdfog_kam['PROJ_AGE_1'].isnull()
        ]
        
        

        df_rpdfog_kam_s1['REDUCTION_FACTOR_S1'] = np.where(
            df_rpdfog_kam_s1['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdfog_kam_s1['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        )      
        
        df_rpdfog_kam_s1['THLB_AREA_DECREASE'] = df_rpdfog_kam_s1['THLB_AREA'] * df_rpdfog_kam_s1['REDUCTION_FACTOR_S1']
        
        df_rpdfog_kam_s1_sum = df_rpdfog_kam_s1.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdfog_kam_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        
            ## scenario-2 ##
        df_rpdfog_kam_s2 = df_rpdfog_kam[
            (df_rpdfog_kam['PROJ_AGE_1'] >= 60) | df_rpdfog_kam['PROJ_AGE_1'].isnull()
        ]
        
        df_rpdfog_kam_s2['REDUCTION_FACTOR_S2'] = np.where(
            df_rpdfog_kam_s2['OVERLAP_TYPE'] != 'IDF only', 
            1, 
            np.where(df_rpdfog_kam_s2['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        )      
        
        df_rpdfog_kam_s2['THLB_AREA_DECREASE'] = df_rpdfog_kam_s2['THLB_AREA'] * df_rpdfog_kam_s2['REDUCTION_FACTOR_S2']
        
        df_rpdfog_kam_s2_sum = df_rpdfog_kam_s2.groupby(['TSA_NAME', 'OVERLAP_TYPE'])[['THLB_AREA_DECREASE']].sum().reset_index()
        
        df_rpdfog_kam_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')  
        
        df_rpdfog_kam_fn = pd.concat([df_rpdfog_kam_s1_sum, df_rpdfog_kam_s2_sum], ignore_index=True)
        
        
        df_rpdfog_fn = pd.concat([df_rpdfog_okn_fn, df_rpdfog_kam_fn], ignore_index=True)
        
 
    except Exception as e:
        raise Exception(f"Error occurred: {e}")  
    
    finally: 
        Duckdb.disconnect_db()
        

    print ('\n Export summary tables') 
    dfs= [df_rpog_sum, df_rpdf_fnl, df_rpdfog_fn]
    datetime= datetime.now().strftime("%Y%m%d_%H%M")
    outfile= os.path.join(wks, 'outputs', f'{datetime}_summary_tables_total_THLB_decrease_v2_details.xlsx')
    write_dfs_to_excel(dfs, outfile)
    


    finish_t = timeit.default_timer() #finish time
    t_sec = round(finish_t-start_t)
    mins = int (t_sec/60)
    secs = int (t_sec%60)
    print (f'\nProcessing Completed in {mins} minutes and {secs} seconds')  
          
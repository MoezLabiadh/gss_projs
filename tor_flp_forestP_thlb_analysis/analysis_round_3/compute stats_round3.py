import warnings
warnings.simplefilter(action='ignore')

import os
import timeit
import duckdb
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from datetime import datetime

class DuckDBConnector:
    def __init__(self, db=':memory:'):
        self.db = db
        self.conn = None
    
    def connect_to_db(self):
        """Connects to a DuckDB database and installs spatial extension."""
        self.conn = duckdb.connect(self.db)
        self.conn.install_extension('spatial')
        self.conn.load_extension('spatial')
        return self.conn
    
    def disconnect_db(self):
        """Disconnects from the DuckDB database."""
        if self.conn is not None:
            self.conn.close()
            self.conn = None
 

def generate_report (workspace, df_list, sheet_list,filename):
    """ Exports dataframes to multi-tab excel spreasheet"""
    outfile= os.path.join(workspace, filename + '.xlsx')

    writer = pd.ExcelWriter(outfile,engine='xlsxwriter')

    for dataframe, sheet in zip(df_list, sheet_list):
        dataframe = dataframe.reset_index(drop=True)
        dataframe.index = dataframe.index + 1

        dataframe.to_excel(writer, sheet_name=sheet, index=False, startrow=0 , startcol=0)

        worksheet = writer.sheets[sheet]
        #workbook = writer.book

        worksheet.set_column(0, dataframe.shape[1], 23)

        col_names = [{'header': col_name} for col_name in dataframe.columns[1:-1]]
        col_names.insert(0,{'header' : dataframe.columns[0], 'total_string': 'Total'})
        col_names.append ({'header' : dataframe.columns[-1], 'total_function': 'sum'})


        worksheet.add_table(0, 0, dataframe.shape[0]+1, dataframe.shape[1]-1, {
            'total_row': True,
            'columns': col_names})

    #writer.save()
    writer.close()
    

def export_dfs_to_sheets(dataframes, sheet_names, filename):
    # Create a new Excel workbook
    wb = Workbook()

    # Remove the default sheet created with the workbook
    wb.remove(wb.active)

    for df, sheet_name in zip(dataframes, sheet_names):
        # Add a new sheet for each dataframe
        ws = wb.create_sheet(title=sheet_name)

        # Write the dataframe to the Excel sheet, ignoring the index
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(filename)  
            

if __name__ == "__main__":
    start_t = timeit.default_timer() #start time 
    
    wks= r'W:\lwbc\visr\Workarea\moez_labiadh\WORKSPACE_2024\20240819_flp_to_thlb_analysis'

    print ('Connecting to databases')    
    print ('..connect to Duckdb') 
    projDB= os.path.join(wks, 'inputs', 'tor_flp_thlb_analysis.db')
    Duckdb= DuckDBConnector(db= projDB)
    Duckdb.connect_to_db()
    dckCnx= Duckdb.conn 
    dckCnx.execute("SET GLOBAL pandas_analyze_sample=1000000")
    
    try:
        '''
        print ('\nCompute Total THLB summary')
        df_tot= dckCnx.execute("""SELECT* EXCLUDE geometry FROM r3_vri_thlb""").df()
        
        bins = [0, 79, 139, 249, 400, float('inf')]
        labels = ['0-79', '80-139', '140-249', '250-400', '400+']
        df_tot['AGE_CLASS'] = pd.cut(df_tot['PROJ_AGE_1'], bins=bins, labels=labels, right=True)
        
        df_tot['CURRENT_THLB_HA'] = df_tot['AREA_HA'] * df_tot['thlb_fact']
        
        df_tot['CURRENT_THLB_GROWING_STOCK_M3'] = df_tot['CURRENT_THLB_HA'] * df_tot['LIVE_STAND_VOLUME_125']
        
        df_tot_sum = df_tot.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['CURRENT_THLB_HA', 'CURRENT_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_tot_sum= df_tot_sum.sort_values(by=['TSA_NAME','BEC_ZONE_CODE'])

        '''

        print ('\nCompute IDF summaries')
        
        df_idf= dckCnx.execute("""SELECT* EXCLUDE geometry FROM r3_idf_vri_thlb WHERE BEC_ZONE_CODE='IDF'""").df()

        df_idf = df_idf.rename(columns={"TSA_NUMBER_DESCRIPTION": "TSA_NAME"})
        
        df_idf['PROJ_AGE_1'] = df_idf['PROJ_AGE_1'].replace(-999, np.nan)
        
        bins = [0, 79, 139, 249, 400, float('inf')]
        labels = ['0-79', '80-139', '140-249', '250-400', '400+']
        df_idf['AGE_CLASS'] = pd.cut(df_idf['PROJ_AGE_1'], bins=bins, labels=labels, right=True)
        
        df_idf['FBP_THLB_HA'] = df_idf['AREA_HA'] * df_idf['thlb_fact']
        
        df_idf['CURRENT_THLB_HA']= df_idf['AREA_HA'] * df_idf['thlb_fact']
        
        df_idf['CURRENT_THLB_GROWING_STOCK_M3'] = df_idf['CURRENT_THLB_HA'] * df_idf['LIVE_STAND_VOLUME_125']
        

        ####### Okanagan scenarios #######
        df_idf_okn= df_idf[df_idf['TSA_NAME']=='Okanagan TSA']
        
        df_idf_okn = df_idf_okn[~df_idf_okn['BEC_SUBZONE'].isin(['mm', 'mw'])]
        
            ## scenario-1 ##
        df_idf_okn_s1 = df_idf_okn[df_idf_okn['PROJ_AGE_1'] >= 100]
        
        
        df_idf_okn_s1['IDF_REDUCTION_FACTOR_S1'] = np.where(
            df_idf_okn_s1['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)

        
        df_idf_okn_s1['FBP_THLB_HA'] = df_idf_okn_s1['CURRENT_THLB_HA'] * df_idf_okn_s1['IDF_REDUCTION_FACTOR_S1']
        
        df_idf_okn_s1['FBP_THLB_GROWING_STOCK_M3'] = df_idf_okn_s1['FBP_THLB_HA'] * df_idf_okn_s1['LIVE_STAND_VOLUME_125']
        
        df_idf_okn_s1_sum = df_idf_okn_s1.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['CURRENT_THLB_HA','CURRENT_THLB_GROWING_STOCK_M3','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_idf_okn_s1_sum['CHANGE_GROWING_STOCK_M3'] = df_idf_okn_s1_sum['CURRENT_THLB_GROWING_STOCK_M3'] - df_idf_okn_s1_sum['FBP_THLB_GROWING_STOCK_M3']

        
        df_idf_okn_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        


            ## scenario-2 ##
        df_idf_okn_s2 = df_idf_okn[df_idf_okn['PROJ_AGE_1'] >= 60]
        
        
        df_idf_okn_s2['IDF_REDUCTION_FACTOR_S2'] = np.where(
            df_idf_okn_s2['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)

        
        df_idf_okn_s2['FBP_THLB_HA'] = df_idf_okn_s2['CURRENT_THLB_HA'] * df_idf_okn_s2['IDF_REDUCTION_FACTOR_S2']
        
        df_idf_okn_s2['FBP_THLB_GROWING_STOCK_M3'] = df_idf_okn_s2['FBP_THLB_HA'] * df_idf_okn_s2['LIVE_STAND_VOLUME_125']
        
        df_idf_okn_s2_sum = df_idf_okn_s2.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['CURRENT_THLB_HA','CURRENT_THLB_GROWING_STOCK_M3','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_idf_okn_s2_sum['CHANGE_GROWING_STOCK_M3'] = df_idf_okn_s2_sum['CURRENT_THLB_GROWING_STOCK_M3'] - df_idf_okn_s2_sum['FBP_THLB_GROWING_STOCK_M3']

        
        df_idf_okn_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_idf_okn_fnl = pd.concat([df_idf_okn_s1_sum, df_idf_okn_s2_sum], ignore_index=True)

          
  
        ####### Kamloops scenarios #######
        df_idf_kam= df_idf[df_idf['TSA_NAME']=='Kamloops TSA']
        
        df_idf_kam= df_idf_kam[~df_idf_kam['BEC_SUBZONE'].isin(['mm', 'mw'])]
        
            ## scenario-1 ##
        df_idf_kam_s1= df_idf_kam[df_idf_kam['PROJ_AGE_1'] >= 100]

        df_idf_kam_s1['IDF_REDUCTION_FACTOR_S1'] = np.where(
            df_idf_kam_s1['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        
        df_idf_kam_s1['FBP_THLB_HA'] = df_idf_kam_s1['CURRENT_THLB_HA'] * df_idf_kam_s1['IDF_REDUCTION_FACTOR_S1']
        
        df_idf_kam_s1['FBP_THLB_GROWING_STOCK_M3'] = df_idf_kam_s1['FBP_THLB_HA'] * df_idf_kam_s1['LIVE_STAND_VOLUME_125']
        
        df_idf_kam_s1_sum = df_idf_kam_s1.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['CURRENT_THLB_HA','CURRENT_THLB_GROWING_STOCK_M3','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_idf_kam_s1_sum['CHANGE_GROWING_STOCK_M3'] = df_idf_kam_s1_sum['CURRENT_THLB_GROWING_STOCK_M3'] - df_idf_kam_s1_sum['FBP_THLB_GROWING_STOCK_M3']

        
        df_idf_kam_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        

            ## scenario-2 ##
        df_idf_kam_s2= df_idf_kam[df_idf_kam['PROJ_AGE_1'] >= 60]

        df_idf_kam_s2['IDF_REDUCTION_FACTOR_S2'] = np.where(
            df_idf_kam_s2['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        
        df_idf_kam_s2['FBP_THLB_HA'] = df_idf_kam_s2['CURRENT_THLB_HA'] * df_idf_kam_s2['IDF_REDUCTION_FACTOR_S2']
        
        df_idf_kam_s2['FBP_THLB_GROWING_STOCK_M3'] = df_idf_kam_s2['FBP_THLB_HA'] * df_idf_kam_s2['LIVE_STAND_VOLUME_125']
        
        df_idf_kam_s2_sum = df_idf_kam_s2.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['CURRENT_THLB_HA','CURRENT_THLB_GROWING_STOCK_M3','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_idf_kam_s2_sum['CHANGE_GROWING_STOCK_M3'] = df_idf_kam_s2_sum['CURRENT_THLB_GROWING_STOCK_M3'] - df_idf_kam_s2_sum['FBP_THLB_GROWING_STOCK_M3']

        
        df_idf_kam_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_idf_kam_fnl = pd.concat([df_idf_kam_s1_sum, df_idf_kam_s2_sum], ignore_index=True)
        
        
        df_idf_fn = pd.concat([df_idf_okn_fnl, df_idf_kam_fnl], ignore_index=True)

        col_order= ['TSA_NAME', 'SCENARIO', 'AGE_CLASS', 'BEC_ZONE_CODE', 'CURRENT_THLB_HA', 'FBP_THLB_HA',
                    'CURRENT_THLB_GROWING_STOCK_M3', 'FBP_THLB_GROWING_STOCK_M3', 'CHANGE_GROWING_STOCK_M3']

        df_idf_fn = df_idf_fn[col_order]

        
        print ('\nCompute Riparian summary')
        
        df_rip= dckCnx.execute("""SELECT* EXCLUDE geometry FROM r3_rip_vri_thlb""").df()
        
        bins = [0, 79, 139, 249, 400, float('inf')]
        labels = ['0-79', '80-139', '140-249', '250-400', '400+']
        df_rip['AGE_CLASS'] = pd.cut(df_rip['PROJ_AGE_1'], bins=bins, labels=labels, right=True)

        
        df_rip['FBP_THLB_HA'] = df_rip['AREA_HA'] * df_rip['thlb_fact']   
        
        
        df_rip['AGE_CLASS'] = df_rip['AGE_CLASS'].cat.add_categories("NO_DATA")
        df_rip.loc[df_rip['AGE_CLASS'].isnull(), ['AGE_CLASS', 'LIVE_STAND_VOLUME_125']] = ["NO_DATA", 0]
        
        
        df_rip['FBP_THLB_GROWING_STOCK_M3'] = df_rip['FBP_THLB_HA'] * df_rip['LIVE_STAND_VOLUME_125']
                
        df_rip_sum = df_rip.groupby(
                ['TSA_NAME', 'AGE_CLASS', 'BEC_ZONE_CODE']
                )[['FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']].sum().reset_index()
        
        df_rip_sum= df_rip_sum.sort_values(by=['TSA_NAME','BEC_ZONE_CODE'])

        
        df_rip_sum.insert(1, 'SCENARIO', 'Riparian')
        
        
        
        ########### FINAL RESULTS ###################
        df_rslt= pd.concat([df_idf_fn, df_rip_sum], ignore_index=True)
        df_rslt = df_rslt.round(2)
        
        
        
        '''
        ########### RAW DATA ###################
        df_rw_idf= pd.concat([df_idf_okn_s1, df_idf_okn_s2, df_idf_kam_s1, df_idf_kam_s2], ignore_index=True)
        
        cols= ['TSA_NAME', 'thlb_fact', 'BEC_ZONE_CODE', 'BEC_SUBZONE', 'PROJ_AGE_1', 'AGE_CLASS',
               'LIVE_STAND_VOLUME_125', 'MDWR_OVERLAP', 'AREA_HA', 'CURRENT_THLB_HA',
               'CURRENT_THLB_GROWING_STOCK_M3', 'IDF_REDUCTION_FACTOR_S1', 
               'IDF_REDUCTION_FACTOR_S2','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']
        
        df_rw_idf = df_rw_idf[cols]
        
        cols_rip= ['TSA_NAME', 'thlb_fact', 'BEC_ZONE_CODE', 'BEC_SUBZONE', 'PROJ_AGE_1', 'AGE_CLASS',
               'LIVE_STAND_VOLUME_125', 'AREA_HA','FBP_THLB_HA', 'FBP_THLB_GROWING_STOCK_M3']
        
        df_rw_rip = df_rip[cols_rip]
        '''
        
   
    except Exception as e:
        raise Exception(f"Error occurred: {e}")  
    
    finally: 
        Duckdb.disconnect_db()
        
 
    
    print ('\n Export summary tables') 
    dfs= [df_rslt]
    sheets= ['vri_summary']
    
    outfile= os.path.join(wks, 'outputs')
    datetime= datetime.now().strftime("%Y%m%d_%H%M")
    generate_report (outfile, dfs, sheets, f'{datetime}_summary_growing_stock.xlsx')
    

    ''' 
    print ('\n Export raw datasets') 
    datetime= datetime.now().strftime("%Y%m%d_%H%M")
    df_rw_idf.to_csv(os.path.join(wks, 'outputs', f'{datetime}_raw_data_growing_stock_IDF.csv'), index=False)
    df_rw_rip.to_csv(os.path.join(wks, 'outputs', f'{datetime}_raw_data_growing_stock_RIPARIAN.csv'), index=False)
    '''
 

    finish_t = timeit.default_timer() #finish time
    t_sec = round(finish_t-start_t)
    mins = int (t_sec/60)
    secs = int (t_sec%60)
    print (f'\nProcessing Completed in {mins} minutes and {secs} seconds')  
        
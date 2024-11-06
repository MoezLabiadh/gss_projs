import warnings
warnings.simplefilter(action='ignore')

import os
import timeit
import duckdb
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from datetime import datetime

class DuckDBConnector:
    def __init__(self, db=':memory:'):
        self.db = db
        self.conn = None
    
    def connect_to_db(self):
        """Connects to a DuckDB database and installs spatial extension."""
        self.conn = duckdb.connect(self.db)
        self.conn.install_extension('spatial')
        self.conn.load_extension('spatial')
        return self.conn
    
    def disconnect_db(self):
        """Disconnects from the DuckDB database."""
        if self.conn is not None:
            self.conn.close()
            self.conn = None
 

def write_dfs_to_excel(dataframes, filename):
    # Create a new Excel workbook and activate the first sheet
    wb = Workbook()
    ws = wb.active

    current_row = 1
    
    for idx, df in enumerate(dataframes):
        # Round all numeric values to 1 decimal place
        df = df.round(2)

        # Write the dataframe to the Excel sheet, ignoring the index
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=current_row):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Define table range
        table_ref = f"A{current_row}:"
        last_column_letter = chr(64 + len(df.columns))  # Get last column's letter
        table_ref += f"{last_column_letter}{current_row + len(df)}"
        
        # Add table layout with a default style
        table = Table(displayName=f"Table{idx+1}", ref=table_ref)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
        
        # Move to the next block of rows for the next dataframe (4-row space)
        current_row += len(df) + 4  # 1 row for the header + df rows + 4 empty rows
    
    # Set column widths: first 4 columns to 14, the rest to 24.5
    for col_idx, col in enumerate(ws.columns, start=1):
        width = 14 if col_idx <= 4 else 24.5
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Save the workbook
    wb.save(filename)
    

def export_dfs_to_sheets(dataframes, sheet_names, filename):
    # Create a new Excel workbook
    wb = Workbook()

    # Remove the default sheet created with the workbook
    wb.remove(wb.active)

    for df, sheet_name in zip(dataframes, sheet_names):
        # Add a new sheet for each dataframe
        ws = wb.create_sheet(title=sheet_name)

        # Write the dataframe to the Excel sheet, ignoring the index
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(filename)  
            

if __name__ == "__main__":
    start_t = timeit.default_timer() #start time 
    
    wks= r'W:\lwbc\visr\Workarea\moez_labiadh\WORKSPACE_2024\20240819_flp_to_thlb_analysis'

    print ('Connecting to databases')    
    print ('..connect to Duckdb') 
    projDB= os.path.join(wks, 'inputs', 'tor_flp_thlb_analysis.db')
    Duckdb= DuckDBConnector(db= projDB)
    Duckdb.connect_to_db()
    dckCnx= Duckdb.conn 
    dckCnx.execute("SET GLOBAL pandas_analyze_sample=1000000")
    
    try:

        print ('\nCompute Gross THLB summaries')
        # thlb by TSA (whole tsa)
        df_tlhb_tsa= dckCnx.execute("""SELECT* EXCLUDE geometry FROM thlb""").df()
        df_tlhb_tsa= df_tlhb_tsa.rename(columns={'tsa_number_description': 'TSA_NAME',
                                                 'thlb_area_ha': 'THLB_AREA'})
        df_tlhb_tsa_sum = df_tlhb_tsa.groupby(['TSA_NAME'])[['THLB_AREA']].sum().reset_index().rename(columns={'THLB_AREA': 'TSA_THLB_AREA'})
        

        
        # thlb by TSA (in plan area)
        df_tlhb_qs= dckCnx.execute("""SELECT*  EXCLUDE geometry FROM thlb_tsa_qs""").df() 
        df_tlhb_qs['QS_THLB_AREA']= df_tlhb_qs['AREA_HA'] * df_tlhb_qs['thlb_fact']
        df_tlhb_qs_sum = df_tlhb_qs.groupby(['TSA_NAME'])[['QS_THLB_AREA']].sum().reset_index()
        
        df_tlhb_qs_sum['RIP_ADJUST_FACTOR'] = np.where(
                df_tlhb_qs_sum['TSA_NAME'] == '100 Mile House TSA', 0.02,
                np.where(df_tlhb_qs_sum['TSA_NAME'] == 'Okanagan TSA', 0.13, 0)
        )
        
        df_tlhb_qs_sum['QS_THLB_AREA_RIP_ADJUSTED'] = df_tlhb_qs_sum['QS_THLB_AREA'] + (
            df_tlhb_qs_sum['QS_THLB_AREA'] * df_tlhb_qs_sum['RIP_ADJUST_FACTOR']
        )
        
        
        df_tlhb_sumAll= pd.merge(df_tlhb_tsa_sum, df_tlhb_qs_sum, on='TSA_NAME')
        
        
        
        print ('\nCompute OGDA summary')
        
        #df_ogda= dckCnx.execute("""SELECT* EXCLUDE geometry FROM ogda_thlb_tsa""").df()
        df_ogda= dckCnx.execute(
            """SELECT*
                FROM r2_2_rip_idf_ogda_thlb_mdwr
               WHERE OVERLAP_TYPE in ('OGDA only', 'IDF/OGDA overlap', 'Riparian/OGDA overlap', 'Riparian/IDF/OGDA overlap') 
                
            """
            ).df()
        
        df_ogda = df_ogda.rename(columns={"TSA_NUMBER_DESCRIPTION": "TSA_NAME"})
        
        
        
        df_ogda['OGDA_THLB_AREA']= df_ogda['AREA_HA'] * df_ogda['thlb_fact']
        
        df_ogda_sum = df_ogda.groupby(['TSA_NAME'])[['OGDA_THLB_AREA']].sum().reset_index()
        
        df_tlhb_sumAll_norip= df_tlhb_sumAll[['TSA_NAME', 'TSA_THLB_AREA', 'QS_THLB_AREA']]
        
        df_ogda_fnl= pd.merge(df_tlhb_sumAll_norip, df_ogda_sum, on='TSA_NAME')
        
        df_ogda_fnl['OGDA_REDUCTION_FACTOR']= 1
        df_ogda_fnl['THLB_AREA_DECREASE'] = df_ogda_fnl['OGDA_THLB_AREA'] * df_ogda_fnl['OGDA_REDUCTION_FACTOR']
        df_ogda_fnl['THLB_AREA_DECREASE_%'] = round((df_ogda_fnl['THLB_AREA_DECREASE'] / df_ogda_fnl['QS_THLB_AREA']) * 100, 1)
        df_ogda_fnl['QS_THLB_AREA_REMAINING'] = df_ogda_fnl['QS_THLB_AREA'] - df_ogda_fnl['THLB_AREA_DECREASE']



        print ('\nCompute IDF summaries')
        
        #df_idf= dckCnx.execute("""SELECT* EXCLUDE geometry FROM idf_thlb_tsa_mdwr""").df()
        df_idf= dckCnx.execute(
            """SELECT*
                FROM r2_2_rip_idf_ogda_thlb_mdwr
               WHERE OVERLAP_TYPE in ('IDF only', 'IDF/OGDA overlap', 'Riparian/IDF overlap', 'Riparian/IDF/OGDA overlap') 
                
            """
            ).df()
        
        df_idf = df_idf.rename(columns={"TSA_NUMBER_DESCRIPTION": "TSA_NAME"})
        
        df_idf['IDF_THLB_AREA']= df_idf['AREA_HA'] * df_idf['thlb_fact']
        

        ####### 100 Mile House scenarios #######
        df_idf_omh= df_idf[df_idf['TSA_NAME']=='100 Mile House TSA']
        df_idf_omh['IDF_REDUCTION_FACTOR']= 0.5
        
        bec_excl= ['mm', 'mw', 'dk', 'xh', 'xm', 'dw', 'xw', 'ww']
        df_idf_omh = df_idf_omh[~df_idf_omh['BEC_SUBZONE'].isin(bec_excl)]

        


        ####### Okanagan scenarios #######
        df_idf_okn= df_idf[df_idf['TSA_NAME']=='Okanagan TSA']
        
        df_idf_okn = df_idf_okn[~df_idf_okn['BEC_SUBZONE'].isin(['mm', 'mw'])]
        
            ## scenario-1 ##
        df_idf_okn_s1 = df_idf_okn[df_idf_okn['PROJ_AGE_1'] >= 100]
        
        #df_idf_okn_s1['IDF_REDUCTION_FACTOR_S1'] = 0
        
        df_idf_okn_s1['IDF_REDUCTION_FACTOR_S1'] = np.where(
            df_idf_okn_s1['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)

        
        df_idf_okn_s1['THLB_AREA_DECREASE'] = df_idf_okn_s1['IDF_THLB_AREA'] * df_idf_okn_s1['IDF_REDUCTION_FACTOR_S1']
        
        df_idf_okn_s1_sum = df_idf_okn_s1.groupby(['TSA_NAME'])[['IDF_THLB_AREA', 'THLB_AREA_DECREASE']].sum().reset_index()
        
        df_idf_okn_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        
        df_tlhb_sumAll_norip= df_tlhb_sumAll[['TSA_NAME', 'TSA_THLB_AREA', 'QS_THLB_AREA']]
        
        df_idf_okn_s1_fnl= pd.merge(df_tlhb_sumAll_norip, df_idf_okn_s1_sum, on='TSA_NAME')
        
        df_idf_okn_s1_fnl['THLB_AREA_DECREASE_%'] = round((df_idf_okn_s1_fnl['THLB_AREA_DECREASE'] / df_idf_okn_s1_fnl['QS_THLB_AREA']) * 100, 1)
        df_idf_okn_s1_fnl['QS_THLB_AREA_REMAINING'] = df_idf_okn_s1_fnl['QS_THLB_AREA'] - df_idf_okn_s1_fnl['THLB_AREA_DECREASE']
        


            ## scenario-2 ##
        df_idf_okn_s2 = df_idf_okn[df_idf_okn['PROJ_AGE_1'] >= 60]
        
        #df_idf_okn_s2['IDF_REDUCTION_FACTOR_S2'] = 0
        
        df_idf_okn_s2['IDF_REDUCTION_FACTOR_S2'] = np.where(
            df_idf_okn_s2['BEC_SUBZONE'].isin(['dk', 'dm']), 0.14, 0.5)
        
        df_idf_okn_s2['THLB_AREA_DECREASE'] = df_idf_okn_s2['IDF_THLB_AREA'] * df_idf_okn_s2['IDF_REDUCTION_FACTOR_S2']
        
        df_idf_okn_s2_sum = df_idf_okn_s2.groupby(['TSA_NAME'])[['IDF_THLB_AREA', 'THLB_AREA_DECREASE']].sum().reset_index()
        
        df_idf_okn_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_tlhb_sumAll_norip= df_tlhb_sumAll[['TSA_NAME', 'TSA_THLB_AREA', 'QS_THLB_AREA']]
        
        df_idf_okn_s2_fnl= pd.merge(df_tlhb_sumAll_norip, df_idf_okn_s2_sum, on='TSA_NAME')
        
        df_idf_okn_s2_fnl['THLB_AREA_DECREASE_%'] = round((df_idf_okn_s2_fnl['THLB_AREA_DECREASE'] / df_idf_okn_s2_fnl['QS_THLB_AREA']) * 100, 1)
        df_idf_okn_s2_fnl['QS_THLB_AREA_REMAINING'] = df_idf_okn_s2_fnl['QS_THLB_AREA'] - df_idf_okn_s2_fnl['THLB_AREA_DECREASE']
        
        df_idf_okn_fnl = pd.concat([df_idf_okn_s1_fnl, df_idf_okn_s2_fnl], ignore_index=True)
        
        
        
          
        
        ####### Kamloops scenarios #######
        df_idf_kam= df_idf[df_idf['TSA_NAME']=='Kamloops TSA']
        
        df_idf_kam= df_idf_kam[~df_idf_kam['BEC_SUBZONE'].isin(['mm', 'mw'])]
        
            ## scenario-1 ##
        df_idf_kam_s1= df_idf_kam[df_idf_kam['PROJ_AGE_1'] >= 100]

        df_idf_kam_s1['IDF_REDUCTION_FACTOR_S1'] = np.where(
            df_idf_kam_s1['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        
        df_idf_kam_s1['THLB_AREA_DECREASE'] = df_idf_kam_s1['IDF_THLB_AREA'] * df_idf_kam_s1['IDF_REDUCTION_FACTOR_S1']
        
        df_idf_kam_s1_sum = df_idf_kam_s1.groupby(['TSA_NAME'])[['IDF_THLB_AREA', 'THLB_AREA_DECREASE']].sum().reset_index()
        
        df_idf_kam_s1_sum.insert(1, 'SCENARIO', 'Scenario 1')
        
        df_tlhb_sumAll_norip= df_tlhb_sumAll[['TSA_NAME', 'TSA_THLB_AREA', 'QS_THLB_AREA']]
        
        df_idf_kam_s1_fnl= pd.merge(df_tlhb_sumAll_norip, df_idf_kam_s1_sum, on='TSA_NAME')
        
        df_idf_kam_s1_fnl['THLB_AREA_DECREASE_%'] = round((df_idf_kam_s1_fnl['THLB_AREA_DECREASE'] / df_idf_kam_s1_fnl['QS_THLB_AREA']) * 100, 1)
        df_idf_kam_s1_fnl['QS_THLB_AREA_REMAINING'] = df_idf_kam_s1_fnl['QS_THLB_AREA'] - df_idf_kam_s1_fnl['THLB_AREA_DECREASE']

        

            ## scenario-2 ##
        df_idf_kam_s2= df_idf_kam[df_idf_kam['PROJ_AGE_1'] >= 60]

        df_idf_kam_s2['IDF_REDUCTION_FACTOR_S2'] = np.where(
            df_idf_kam_s2['MDWR_OVERLAP'].notnull(), 0.25, 0.5)
        
        df_idf_kam_s2['THLB_AREA_DECREASE'] = df_idf_kam_s2['IDF_THLB_AREA'] * df_idf_kam_s2['IDF_REDUCTION_FACTOR_S2']
        
        df_idf_kam_s2_sum = df_idf_kam_s2.groupby(['TSA_NAME'])[['IDF_THLB_AREA', 'THLB_AREA_DECREASE']].sum().reset_index()
        
        df_idf_kam_s2_sum.insert(1, 'SCENARIO', 'Scenario 2')
        
        df_tlhb_sumAll_norip= df_tlhb_sumAll[['TSA_NAME', 'TSA_THLB_AREA', 'QS_THLB_AREA']]
        
        df_idf_kam_s2_fnl= pd.merge(df_tlhb_sumAll_norip, df_idf_kam_s2_sum, on='TSA_NAME')
        
        df_idf_kam_s2_fnl['THLB_AREA_DECREASE_%'] = round((df_idf_kam_s2_fnl['THLB_AREA_DECREASE'] / df_idf_kam_s2_fnl['QS_THLB_AREA']) * 100, 1)
        df_idf_kam_s2_fnl['QS_THLB_AREA_REMAINING'] = df_idf_kam_s2_fnl['QS_THLB_AREA'] - df_idf_kam_s2_fnl['THLB_AREA_DECREASE']


        df_idf_kam_fnl = pd.concat([df_idf_kam_s1_fnl, df_idf_kam_s2_fnl], ignore_index=True)
        
        
        
        df_idf_fnl = pd.concat([df_idf_okn_fnl, df_idf_kam_fnl], ignore_index=True)
        
        
        
        print ('\nCompute Riparian summary - FBP')
        
        #df_rip_fbp= dckCnx.execute("""SELECT* EXCLUDE geometry FROM rip_fbp_thlb_tsa""").df()
        df_rip_fbp= dckCnx.execute(
            """SELECT*
                FROM r2_2_rip_idf_ogda_thlb_mdwr
               WHERE OVERLAP_TYPE in ('Riparian only', 'Riparian/IDF overlap', 'Riparian/OGDA overlap', 'Riparian/IDF/OGDA overlap') 
                
            """
            ).df()
        
        df_rip_fbp = df_rip_fbp.rename(columns={"TSA_NUMBER_DESCRIPTION": "TSA_NAME"})
        
        df_rip_fbp['RIP_FBP_THLB_AREA']= df_rip_fbp['AREA_HA'] * df_rip_fbp['thlb_fact']
        
        df_rip_fbp_sum = df_rip_fbp.groupby(['TSA_NAME'])[['RIP_FBP_THLB_AREA']].sum().reset_index()
        
        df_rip_fbp_fnl= pd.merge(df_tlhb_sumAll, df_rip_fbp_sum, on='TSA_NAME')
        
        df_rip_fbp_fnl['RIP_REDUCTION_FACTOR']= 1
        df_rip_fbp_fnl['THLB_AREA_DECREASE'] = df_rip_fbp_fnl['RIP_FBP_THLB_AREA'] * df_rip_fbp_fnl['RIP_REDUCTION_FACTOR']
        df_rip_fbp_fnl['THLB_AREA_DECREASE_%'] = round((df_rip_fbp_fnl['THLB_AREA_DECREASE'] / df_rip_fbp_fnl['QS_THLB_AREA']) * 100, 1)
        df_rip_fbp_fnl['QS_THLB_AREA_REMAINING'] = df_rip_fbp_fnl['QS_THLB_AREA'] - df_rip_fbp_fnl['THLB_AREA_DECREASE']
        
      

        print ('\nCompute Riparian summary - KAM')
        
        df_rip_kam= dckCnx.execute("""SELECT* EXCLUDE geometry FROM rip_kam_thlb""").df()
        
        df_rip_kam['TSA_NAME']= 'Kamloops TSA'
        
        df_rip_kam['RIP_KAM_THLB_AREA']= df_rip_kam['AREA_HA'] * df_rip_kam['thlb_fact']
        
        df_rip_kam_sum = df_rip_kam.groupby(['TSA_NAME'])[['RIP_KAM_THLB_AREA']].sum().reset_index()
        
        
        df_rip_kam_fnl= pd.merge(df_tlhb_sumAll, df_rip_kam_sum, on='TSA_NAME')
        
        df_rip_kam_fnl['RIP_REDUCTION_FACTOR']= 1
        df_rip_kam_fnl['THLB_AREA_DECREASE'] = df_rip_kam_fnl['RIP_KAM_THLB_AREA'] * df_rip_kam_fnl['RIP_REDUCTION_FACTOR']
        df_rip_kam_fnl['THLB_AREA_DECREASE_%'] = round((df_rip_kam_fnl['THLB_AREA_DECREASE'] / df_rip_kam_fnl['QS_THLB_AREA']) * 100, 1)
        df_rip_kam_fnl['QS_THLB_AREA_REMAINING'] = df_rip_kam_fnl['QS_THLB_AREA'] - df_rip_kam_fnl['THLB_AREA_DECREASE']
        
   
    except Exception as e:
        raise Exception(f"Error occurred: {e}")  
    
    finally: 
        Duckdb.disconnect_db()
        

    print ('\n Export summary tables') 
    dfs= [df_idf_fnl, df_ogda_fnl, df_rip_fbp_fnl, df_rip_kam_fnl]
    datetime= datetime.now().strftime("%Y%m%d_%H%M")
    outfile= os.path.join(wks, 'outputs', f'{datetime}_summary_tables.xlsx')
    write_dfs_to_excel(dfs, outfile)
    

    ''' 
    print ('\n Export IDF datasets') 
    dataframes=[df_idf_kam_s1,
                df_idf_kam_s2,
                df_idf_okn_s1,
                df_idf_okn_s2]
    
    sheet_names=['kamloops_scenario1',
                 'kamloops_scenario2',
                 'okanagan_scenario1',
                 'okanagan_scenario2']
    
    outfile_idf= os.path.join(wks, 'outputs', f'{datetime}_idf_data.xlsx')
    export_dfs_to_sheets(dataframes, sheet_names, outfile_idf)
    '''

    finish_t = timeit.default_timer() #finish time
    t_sec = round(finish_t-start_t)
    mins = int (t_sec/60)
    secs = int (t_sec%60)
    print (f'\nProcessing Completed in {mins} minutes and {secs} seconds')  
        